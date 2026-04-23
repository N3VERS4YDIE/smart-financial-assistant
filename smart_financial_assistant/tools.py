"""LangChain tool definitions for SIIGO API calls and XLSX report parsing."""

from __future__ import annotations

import datetime as dt
import json
from io import BytesIO
from typing import Any
from urllib.parse import urlparse

import requests
from langchain_core.tools import tool
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from .config import Settings
from .models import TrialBalanceByThirdPartyRequest, TrialBalanceRequest
from .siigo_client import SiigoClient


def _to_json(data: dict[str, Any]) -> str:
    """Serialize dictionaries to pretty JSON preserving unicode characters."""
    return json.dumps(data, ensure_ascii=False, indent=2)


def _is_allowed_host(url: str, allowed_hosts: tuple[str, ...]) -> bool:
    """Validate that a report URL is HTTPS and belongs to an allowed host."""
    parsed = urlparse(url)
    if parsed.scheme.lower() != "https":
        return False
    host = (parsed.hostname or "").lower()
    return any(host == pattern or host.endswith(f".{pattern}") for pattern in allowed_hosts)


def _safe_cell(value: Any) -> Any:
    """Convert spreadsheet cell values into JSON-safe scalar representations."""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (dt.date, dt.datetime, dt.time)):
        return value.isoformat()
    return str(value)


class ParseXlsxReportRequest(BaseModel):
    """Input schema for downloading and previewing SIIGO XLSX reports."""

    report_url: str = Field(..., description="URL HTTPS del reporte .xlsx generado por SIIGO")
    max_rows_per_sheet: int = Field(
        default=30,
        ge=1,
        le=100,
        description="Numero maximo de filas de vista previa por hoja",
    )


def build_siigo_tools(client: SiigoClient, settings: Settings):
    """Build and return the full set of SIIGO and report-processing LangChain tools."""

    @tool("get_accounts_payable")
    def get_accounts_payable() -> str:
        """Obtiene cuentas por pagar reales desde SIIGO."""
        return _to_json(client.get_accounts_payable())

    @tool("get_trial_balance", args_schema=TrialBalanceRequest)
    def get_trial_balance(
        account_start: str,
        account_end: str,
        year: int,
        month_start: int,
        month_end: int,
        includes_tax_difference: bool = True,
    ) -> str:
        """Consulta balance de prueba real desde SIIGO usando rango de cuentas y periodo."""
        payload = TrialBalanceRequest(
            account_start=account_start,
            account_end=account_end,
            year=year,
            month_start=month_start,
            month_end=month_end,
            includes_tax_difference=includes_tax_difference,
        )
        return _to_json(client.get_trial_balance(payload.model_dump()))

    @tool("get_trial_balance_by_thirdparty", args_schema=TrialBalanceByThirdPartyRequest)
    def get_trial_balance_by_thirdparty(
        account_start: str,
        account_end: str,
        year: int,
        month_start: int,
        month_end: int,
        includes_tax_difference: bool = False,
    ) -> str:
        """Consulta balance de prueba por tercero real desde SIIGO."""
        payload = TrialBalanceByThirdPartyRequest(
            account_start=account_start,
            account_end=account_end,
            year=year,
            month_start=month_start,
            month_end=month_end,
            includes_tax_difference=includes_tax_difference,
        )
        return _to_json(client.get_trial_balance_by_thirdparty(payload.model_dump()))

    @tool("download_and_parse_xlsx_report", args_schema=ParseXlsxReportRequest)
    def download_and_parse_xlsx_report(report_url: str, max_rows_per_sheet: int = 30) -> str:
        """Descarga un reporte .xlsx desde SIIGO/Azure Blob y devuelve una vista estructurada para analisis."""
        if not _is_allowed_host(report_url, settings.allowed_report_hosts):
            raise RuntimeError(
                "URL de reporte no permitida. Verifica host y usa una URL HTTPS de SIIGO/Azure Blob autorizada."
            )

        try:
            response = requests.get(report_url, timeout=settings.report_download_timeout)
            response.raise_for_status()
        except requests.RequestException as exc:
            raise RuntimeError(f"No fue posible descargar el reporte: {exc}") from exc

        max_size_bytes = settings.report_max_size_mb * 1024 * 1024
        if len(response.content) > max_size_bytes:
            raise RuntimeError(
                f"El archivo excede el limite permitido de {settings.report_max_size_mb} MB."
            )

        content_type = response.headers.get("Content-Type", "").lower()
        if "spreadsheetml" not in content_type and not report_url.lower().endswith(".xlsx"):
            raise RuntimeError("El recurso descargado no parece ser un archivo .xlsx valido.")

        try:
            workbook = load_workbook(filename=BytesIO(response.content), data_only=True, read_only=True)
        except Exception as exc:
            raise RuntimeError(f"No fue posible leer el .xlsx descargado: {exc}") from exc

        sheets: list[dict[str, Any]] = []
        max_sheets = 6
        for sheet_name in workbook.sheetnames[:max_sheets]:
            ws = workbook[sheet_name]
            preview_rows: list[list[Any]] = []

            for row in ws.iter_rows(values_only=True):
                serialized = [_safe_cell(cell) for cell in row]
                if any(cell not in (None, "") for cell in serialized):
                    preview_rows.append(serialized)
                if len(preview_rows) >= max_rows_per_sheet:
                    break

            sheets.append(
                {
                    "sheet": sheet_name,
                    "preview_row_count": len(preview_rows),
                    "preview_rows": preview_rows,
                }
            )

        workbook.close()

        return _to_json(
            {
                "report_url": report_url,
                "content_type": content_type,
                "downloaded_bytes": len(response.content),
                "sheet_count": len(workbook.sheetnames),
                "sheet_limit_applied": max_sheets,
                "max_rows_per_sheet": max_rows_per_sheet,
                "sheets": sheets,
            }
        )

    return [
        get_accounts_payable,
        get_trial_balance,
        get_trial_balance_by_thirdparty,
        download_and_parse_xlsx_report,
    ]
